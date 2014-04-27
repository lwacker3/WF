import sys
import re
import os
import csv
from xlrd import open_workbook, empty_cell
from xlwt import Workbook

from openpyxl import Workbook as wb
from openpyxl.cell import get_column_letter



from datetime import date
import logging

from pdfminer.pdfparser import PDFDocument, PDFParser
from pdfminer.pdfinterp import PDFResourceManager, PDFPageInterpreter, process_pdf
from pdfminer.pdfdevice import PDFDevice, TagExtractor
from pdfminer.converter import XMLConverter, HTMLConverter, TextConverter
from pdfminer.cmapdb import CMapDB
from pdfminer.layout import LAParams


	
######################################################################################################
######################################################################################################
######### SETUP GLOBAL VARS 

#real expressions that need to be compiled
re_spaces = re.compile(" +")
re_contract = re.compile("[A-Z][A-Z][A-Z]\d\d")
re_settlement = re.compile("Settlement")
re_avg = re.compile("Avg")
re_curr = re.compile("USD")		
re_stop = re.compile("Total Open Trade Equity")

#necessary constants 
account_0_number = 'PIO79640'
account_1_number = 'PIO79646'
account_2_number = "PIO79647"
account_3_number = "PIO79648"

######################################################################################################
######################################################################################################
######### STATIC HELPER METHODS

month_dictionary = {1:'JAN', 2:'FEB', 3:'MAR', 4:'APR', 5:'MAY', 6:'JUN', 7:'JUL', 8:'AUG', 9:'SEP', 10:'OCT', 11:'NOV', 0:'DEC'}

# static helper methods 

# creates data structure to store the positions found
def create_model(): 
	account_dict = {} 
	account_dict[account_0_number] = {'GOLD' : [], 'SILVER' : [], 'EURO': []}
	account_dict[account_1_number] = {'GOLD' : [], 'SILVER' : [], 'EURO': []}
	account_dict[account_2_number] = {'GOLD' : [], 'SILVER' : [], 'EURO': []}
	return account_dict

# formatting for openpyxl retartedness 
def get_letter_val_openpyxl(_int):

	return get_column_letter(_int)


# removes + sign from positons that are posititve	
def reformat_position_amount(position_amount):
	if '+' in position_amount: 
		return position_amount[1:]
	else: 
		return position_amount


# gets the list of global contracts based on time of year, and the year
def get_global_contracts(): 
	today =  date.today()
	# plus one because the current month's contract is not active 
	month = today.month + 1
	day = today.day
	year = today.year - 2000
	if day < 5: 
		month = today.month
	month_list = []
	iterator = 0
	while iterator < 90:
		month_int = (month + iterator)%12
		if month_int == 1: 
			year +=1 
		month_str = month_dictionary[month_int]
		year_str = str(year)
		month_list.append(month_str + year_str)
		iterator +=1
	return month_list

# gets most recent statement in the directory. 
# warns the user which pdf its reading
def get_recent_pdf(logging_flag): 

	today =  date.today()
	month = today.month
	day = today.day
	year = today.year

	search = True
	#variable to keep track of if a year has already passed. Set this to true so
	# you can't infinitely loop over years. 
	year_passed = False
	while search: 
		day -= 1
		if day == 0:  
			day = 31
			month -=1
			if (month == 0):
				if year_passed == True: 
					raise Exception('neither the pdf ' + file_str_2 + ' or the pdf ' +  file_str_1 +  " exists in this directory- try entering the file you would like to parse manually") 
				year_passed = True
				month = 12

		file_str_1 = str(month)  + "-" + str(day) + ".pdf"
		file_str_2 = str(month) + "-" + str(day)
	
		try:
			f = open(file_str_1, 'r')
			search = False
			f.close()
			if logging_flag:
				logging.warn(" Most current statement found is from " + file_str_2)
			return [file_str_1]
		except:
			try:  
				f = open(file_str_2, 'r')
				search = False
				f.close() 
				if logging_flag:
					logging.warn(" Most current statement found is from " + file_str_2)
				return [file_str_2]
			except: 
				continue

	# this shouldn't happen, but if it does, raise exception
	raise Exception('neither the pdf ' + file_str_2 + ' or the pdf ' +  file_str_1 +  " exists in this directory- try entering the file you would like to parse manually")




######################################################################################################
######################################################################################################
########### CREATE THE MODEL 

global_contracts = get_global_contracts() 


######################################################################################################
######################################################################################################
########## FIND THE POSITIONS

#state-machine style method that scrapes the pdf for the current positions in each account. returns the data structure, just populated properly 
def find_positions(fin):

	account_dict = create_model()

	if fin == None: 
		raise ValueError("there is no text file to read from")
        
	fin = open(fin, "r")
	data_list = fin.readlines()
	fin.close()
	line = 0 
	
	#state variables: 
	contract_search = True # the state in which you are looking for a new contract, this only gets activated AFTER we pass over P79640 account 
	settlement_search =  False # the state in which you are looking for a net position in a specific contract 
	con_type = "None"
	con_str = "None"
	search_content = False
	acct_change = False
	
	current_account = account_0_number
	while (line < len(data_list)):
		next_line = data_list[line]


		# Found an account, searching through contracts - need to identify this account
		if(contract_search): 

			# we have found the 0th account.  This is the first identifcation of the program
			if re.search(account_0_number, next_line): 
					search_content  = False
					current_account = account_0_number

			# we have found the first account
			elif (acct_change == False and re.search(account_1_number, next_line)): 
				search_content = False
				acct_change = True
				current_account = account_1_number

			#this only can occur in the contract searching state. 
			elif (re.search(account_2_number, next_line) and acct_change == False):
					current_account = account_2_number
					acct_change = True	
					search_content  = False

			elif (acct_change == False and re.search(account_3_number, next_line)): 
				current_account = account_3_number
				search_content = False
				acct_change= True
				# refactor this
				line = 1000000000000000
				
			#this decides when the program should start searching for positions. 		
			elif (re.search(" P O S I T I O N S ", next_line)): 
				if current_account == account_0_number: 
					logging.warn('Warning! You have a position in the P79640 Account!')
				search_content = True		
				
			#searching for positions 
			elif (search_content): 		
				contract = re_contract.search(next_line)
				currency = re_curr.search(next_line)
				#line to stop the program from searching non-open positions in the pdf 
				if (re.search("Total Margin Call", next_line) and acct_change == True): 
					acct_change = False
					search_content = False
		
				if (currency != None and contract != None) : 
					
					contract_str = contract.group()
			
					if re.search("GOLD", next_line):
							con_type = "GOLD"
					elif re.search("SILVER", next_line):
							con_type = "SILVER"
					elif re.search("IMM EURO", next_line):
							con_type = "EURO"
					else:
							raise ValueError("You have some strange contract type that is neither Gold, Siler or Eurodollars in your account %s" % current_account)
					settlement_search  = True
					contract_search  = False 
		
		# "Seetlement search state? " - actually getting the long or short positions - con_type is passed in... 
		elif (settlement_search):
			settlement_found=  (re_settlement.search(next_line))
			avg_found  = re_avg.search(next_line)
			if (avg_found != None or settlement_found != None): 
				terms_list = re_spaces.split(next_line)
				long_pos = terms_list[1]
				short_pos = terms_list[2]
				re_pos = re.compile('\d+')
				if long_pos == "*":
						   num_pos = short_pos
						   sign_str= "-"
				if short_pos == "*":
						   num_pos = long_pos
						   sign_str = "+"
				formatted_pos = re_pos.search(num_pos)
				pos_str = sign_str+ formatted_pos.group()
				
				# checking for mishaps - silver in the gold account, gold in the silver account 
				if (current_account == account_1_number and con_type=='SILVER'):
					logging.warn("AHHH You bone head!")
					logging.warn("You have a position of " + contract_str + " " + pos_str + " silver in the Gold %s account!" % account_1_number)
				elif (current_account == account_2_number and con_type=="GOLD"): 
					logging.warn("Asleep at the trading desk again?")
					logging.warn("You have a position of " + contract_str + " " + pos_str + " gold in the Silver %s account!" % account_2_number)


				contract_search = True
				settlement_search = False
			
				# update the model
				account_dict[current_account][con_type].append((contract_str, con_type, pos_str))
				
		line +=1

	return account_dict

######################################################################################################
######################################################################################################
############ FUNCTIONS THAT CREATE THE EXCEL DOCUMENT

def create_trading_sheet(positions, global_workbook):
	#trade_sheet = global_workbook.add_sheet('Formatted Positions', cell_overwrite_ok=True)
	trade_sheet = global_workbook.create_sheet()
	trade_sheet.title = "Formatted Positions"

	#trade_sheet = global_workbook.add_sheet('Formatted Positions', cell_overwrite_ok=True)
	trade_sheet.cell('%s%s' % ('A',1)).value = 'Contract'
	trade_sheet.cell('%s%s' % ('B',1)).value = 'GOLD 646'
	trade_sheet.cell('%s%s' % ('C',1)).value = 'EURO 646'
	trade_sheet.cell('%s%s' % ('D',1)).value = 'SILV 647'
	trade_sheet.cell('%s%s' % ('E',1)).value = 'EURO 647'


	# trade_sheet.write('A', 1, "Contract")
	# trade_sheet.write('B', 2, "GOLD, 646")
	# trade_sheet.write('C', 3, "EURO, 646")
	# trade_sheet.write('D', 4, "SILV, 647")
	# trade_sheet.write('E', 5, "EURO, 647")
	
	def contract_iterate(positions, column_number):

		trade_sheet.cell('%s%s' % ('A',1)).value = 'Contract'
		column_letter = get_letter_val_openpyxl(column_number)
		pos_counter = 0 
		global_counter = 1
		position_length = len(positions)
	#loop to iterate over the positions
		while (global_counter < len(global_contracts)):
			if pos_counter < position_length:
				position = positions[pos_counter]
				pos_contract = position[0]
				pos_amount = position[2]
				contract = global_contracts[global_counter]

				if (contract== pos_contract):
				#write the contract label col. (+1 because of the heading) 
					trade_sheet.cell('%s%s' % ('A', global_counter+1)).value = contract

					#trade_sheet.write(global_counter+1, 0, contract)
				# write the contract value 
					trade_sheet.cell('%s%s' % (column_letter, global_counter + 1)).value=reformat_position_amount(pos_amount)
					#trade_sheet.row(global_counter+1).set_cell_number(column_number, pos_amount)
				#position is next in line, iterate both counters
					pos_counter +=1
					global_counter +=1
				else: 
					trade_sheet.cell('%s%s' % ('A', global_counter + 1 )).value=global_contracts[global_counter]
					trade_sheet.cell('%s%s' % (column_letter, global_counter + 1)).value=0
					#trade_sheet.write(global_counter+1, 0, global_contracts[global_counter])
					#trade_sheet.row(global_counter+1).set_cell_number(column_number, 0)
					global_counter +=1

			else:
				while (global_counter < len(global_contracts)): 	
					trade_sheet.cell('%s%s' % ('A', global_counter+1)).value = global_contracts[global_counter]
					trade_sheet.cell('%s%s' % (column_letter, global_counter + 1)).value= 0 

					#trade_sheet.write(global_counter+1, 0, global_contracts[global_counter])
					#trade_sheet.row(global_counter+1).set_cell_number(column_number, 0)
					global_counter +=1

	gold_positions = positions[account_1_number]['GOLD']
	euro1_positions = positions[account_1_number]['EURO']
	silver_positions = positions[account_2_number]['SILVER']
	euro2_positions = positions[account_2_number]['EURO']
		
	contract_iterate(gold_positions, 2)
	contract_iterate(euro1_positions, 3)
	contract_iterate(silver_positions,4)
	contract_iterate(euro2_positions, 5)

	return global_workbook


def create_reference_sheet(positions, global_workbook): 
		
		ws1 = global_workbook.create_sheet() 
		reference_sheet = global_workbook.create_sheet()
		reference_sheet.title = "Positions by Account"

		# account break down 
		column = 1
		for account, commodities in positions.items(): 
			row = 1
			# writes the acct number
			letter = get_letter_val_openpyxl(column)
			reference_sheet.cell('%s%s'%(letter, 1)).value = account
			
			#reference_sheet.write(0, column, account)
			print column
			print "before 2nd for loop"
			for c_type, pos_list in commodities.items(): 
				letter = get_letter_val_openpyxl(column)
				reference_sheet.cell('%s%s' % (letter,2)).value =  c_type
				#reference_sheet.write(1, column, c_type)
				row=3
				# go down a row 
				if pos_list == {}: 
					column +=2
				
				for position in pos_list:

					letter = get_letter_val_openpyxl(column)
				
					reference_sheet.cell('%s%s' % (letter,row)).value =  position[0]
					#reference_sheet.write(row, column, position[0])
					column +=1 
					letter = get_letter_val_openpyxl(column)

					# if positive position, take out the "+" sign 
					position_amount = reformat_position_amount(position[2])
					reference_sheet.cell('%s%s' % (letter,row)).value = position_amount
					#reference_sheet.write(row, column, position[2])
					column -=1
					row +=1 
				
				column +=2

		return global_workbook

def populate_workbook(positions):
	global_book = wb()
	global_book = create_reference_sheet(positions, global_book)
	global_book = create_trading_sheet(positions, global_book)
	try: 
		os.remove('open_positions.xlsx')
	except: 
		logging.warn('no previous workbook found to delete. Not a problem, if you are running this for the first time in a new folder. ')
	global_book.save('open_positions.xlsx')
	file_name = get_recent_pdf(logging_flag=False)[0]
	parsed_name = file_name[0:-4]

	global_book.save('./Archives/%s.xlsx' % parsed_name )

######################################################################################################
######################################################################################################
############# pdf_to_text method, downloaded from the internet for personal use only in this python file. 

def main(argv):
	import getopt
	def usage():
		print ('usage: %s [-d] [-p pagenos] [-m maxpages] [-P password] [-o output] '
			   '[-n] [-A] [-M char_margin] [-L line_margin] [-W word_margin] [-F boxes_flow] '
			   '[-Y layout_mode] [-O output_dir] [-t text|html|xml|tag] [-c codec] [-s scale] file ...' % argv[0])
		return 100
	try:
		(opts, args) = getopt.getopt(argv[1:], 'dp:m:P:o:nAM:L:W:F:Y:O:t:c:s:')
	except getopt.GetoptError:
		return usage()
	if not args: 
		args = get_recent_pdf(logging_flag=True)
	# debug option
	debug = 0
	# input option
	password = ''
	pagenos = set()
	maxpages = 0
	# output option
	outfile = "text_output.txt"
	outtype = None
	outdir = None
	layoutmode = 'normal'
	codec = 'utf-8'
	pageno = 1
	scale = 1
	showpageno = True
	laparams = LAParams()
	CMapDB.debug = debug
	PDFResourceManager.debug = debug
	PDFDocument.debug = debug
	PDFParser.debug = debug
	PDFPageInterpreter.debug = debug
	PDFDevice.debug = debug

	rsrcmgr = PDFResourceManager()
	if not outtype:
		outtype = 'text'
		if outfile:
			if outfile.endswith('.htm') or outfile.endswith('.html'):
				outtype = 'html'
			elif outfile.endswith('.xml'):
				outtype = 'xml'
			elif outfile.endswith('.tag'):
				outtype = 'tag'
	if outfile:
		outfp = file(outfile, 'w')
	else:
		outfp = sys.stdout
	if outtype == 'text':
		device = TextConverter(rsrcmgr, outfp, codec=codec, laparams=laparams)
	elif outtype == 'xml':
		device = XMLConverter(rsrcmgr, outfp, codec=codec, laparams=laparams, outdir=outdir)
	elif outtype == 'html':
		device = HTMLConverter(rsrcmgr, outfp, codec=codec, scale=scale,
							   layoutmode=layoutmode, laparams=laparams, outdir=outdir)
	elif outtype == 'tag':
		device = TagExtractor(rsrcmgr, outfp, codec=codec)
	else:
		return usage()
	for fname in args:
		fp = file(fname, 'rb')
		process_pdf(rsrcmgr, device, fp, pagenos, maxpages=maxpages, password=password,
					check_extractable=True)
		fp.close()
	device.close()
	outfp.close()
	readfile = outfile
	positions  = find_positions(readfile)
	populate_workbook(positions)
	return

if __name__ == '__main__': sys.exit(main(sys.argv))

